#!/usr/bin/env python3
"""Extract structured header, line and total data from a Bickers quote workbook."""

from __future__ import annotations

import argparse
import json
import math
import re
from pathlib import Path

from openpyxl import load_workbook


NON_PRICES = {"", "tbc", "n/a", "f.o.c", "foc", "production"}


def text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def money(value):
    if isinstance(value, (int, float)) and math.isfinite(value):
        return float(value)
    raw = text(value)
    if raw.lower() in NON_PRICES:
        return None
    cleaned = re.sub(r"[^0-9.]", "", raw)
    if not cleaned:
        return None
    try:
        amount = float(cleaned)
    except ValueError:
        return None
    return -amount if raw.startswith("-") or raw.startswith("(") else amount


def first_numeric(values):
    for value in values:
        parsed = money(value)
        if parsed is not None:
            return parsed
    return None


def extract(path: Path):
    values_book = load_workbook(path, data_only=True, read_only=False)
    formula_book = load_workbook(path, data_only=False, read_only=False)
    values = values_book.worksheets[0]
    formulas = formula_book.worksheets[0]

    header = {
        "quoteDate": text(values["A3"].value),
        "jobNumber": text(values["D3"].value),
        "quoteNumber": text(values["E3"].value),
        "productionCompany": text(values["A5"].value),
        "production": text(values["D5"].value),
        "productionContact": text(values["E5"].value),
        "location": text(values["A7"].value),
        "shootDates": text(values["D7"].value),
        "bickersContact": text(values["E7"].value),
        "serviceDescription": text(values["A9"].value),
    }
    quote_date_formula = text(formulas["A3"].value)
    header["quoteDateReliable"] = "TODAY(" not in quote_date_formula.upper()

    current_section = ""
    lines = []
    total_candidates = []
    max_row = min(max(values.max_row, 1), 600)
    for row in range(10, max_row + 1):
        description = text(values.cell(row, 1).value)
        quantity_raw = values.cell(row, 5).value
        unit_raw = values.cell(row, 6).value
        total_raw = values.cell(row, 7).value
        row_values = [values.cell(row, column).value for column in range(1, 8)]
        if not any(text(value) for value in row_values):
            continue

        row_label = " ".join(text(value) for value in row_values)
        if re.search(r"total price|excludes vat|grand total|total due", row_label, re.I):
            amount = first_numeric(reversed(row_values))
            if amount is not None:
                total_candidates.append({"row": row, "label": row_label, "amount": amount})
            continue

        qty = money(quantity_raw)
        unit_price = money(unit_raw)
        line_total = money(total_raw)
        is_section = bool(description and qty is None and unit_price is None and line_total is None)
        if is_section and not re.match(r"^description$", description, re.I):
            current_section = description
            continue
        if not description and line_total is None:
            continue
        if re.match(r"^description$", description, re.I):
            continue
        if re.match(r"^(notes|terms|excludes vat)", description, re.I):
            continue

        if re.search(r"discount|\bless\b", description, re.I) and line_total is not None:
            line_total = -abs(line_total)
        lines.append({
            "row": row,
            "section": current_section,
            "description": description,
            "quantity": qty if qty is not None else text(quantity_raw),
            "unitPrice": unit_price if unit_price is not None else text(unit_raw),
            "lineTotal": line_total,
            "pricingState": "priced" if line_total is not None else text(total_raw).lower() or "unselected",
        })

    active_lines = [line for line in lines if isinstance(line["lineTotal"], (int, float)) and line["lineTotal"] != 0]
    calculated_total = round(sum(line["lineTotal"] for line in active_lines), 2)
    document_total = total_candidates[-1]["amount"] if total_candidates else calculated_total
    variance = round(document_total - calculated_total, 2)
    issues = []
    if not header["jobNumber"]:
        issues.append("Missing job number")
    if not header["quoteNumber"]:
        issues.append("Missing quote number")
    if not active_lines:
        issues.append("No priced lines")
    if abs(variance) > 0.02:
        issues.append(f"Document total differs from extracted lines by {variance:.2f}")
    confidence = "high" if not issues else ("medium" if len(issues) == 1 else "low")

    return {
        "schemaVersion": 1,
        "source": {"name": path.name, "path": str(path)},
        "sheet": values.title,
        "header": header,
        "lineItems": lines,
        "activeLineCount": len(active_lines),
        "documentTotal": document_total,
        "calculatedTotal": calculated_total,
        "variance": variance,
        "totalCandidates": total_candidates,
        "confidence": {"level": confidence, "issues": issues},
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("paths", nargs="+")
    parser.add_argument("--pretty", action="store_true")
    args = parser.parse_args()
    results = []
    for raw_path in args.paths:
        path = Path(raw_path).expanduser().resolve()
        try:
            results.append(extract(path))
        except Exception as exc:  # Keep a bulk extraction moving and report the exact file.
            results.append({"source": {"name": path.name, "path": str(path)}, "error": str(exc)})
    print(json.dumps(results, indent=2 if args.pretty else None, ensure_ascii=False, default=str))


if __name__ == "__main__":
    main()
